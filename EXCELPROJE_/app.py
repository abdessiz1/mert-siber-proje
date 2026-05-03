from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response
import pandas as pd
import os, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import csv

app = Flask(__name__)
app.secret_key = "siber_master_key_2026"

EXCEL_FILE = 'kayitlar.xlsx'
KOLONLAR = ['ID', 'Unvan', 'Ad', 'Soyad', 'Numara_TC']
GECERLI_UNVANLAR = {'Öğrenci', 'Öğretmen', 'Veli'}

# ─── EXCEL YARDIMCI ────────────────────────────────────────────────────────────

def veritabani_kontrol():
    if not os.path.exists(EXCEL_FILE):
        pd.DataFrame(columns=KOLONLAR).to_excel(EXCEL_FILE, index=False)

def verileri_oku() -> pd.DataFrame:
    if not os.path.exists(EXCEL_FILE):
        veritabani_kontrol()
    try:
        df = pd.read_excel(EXCEL_FILE, dtype=str).fillna("")
        if not df.empty:
            df['ID'] = pd.to_numeric(df['ID'], errors='coerce').fillna(0).astype(int)
        return df
    except Exception:
        return pd.DataFrame(columns=KOLONLAR)

def verileri_yaz(df: pd.DataFrame):
    df = df.copy()
    df['ID'] = pd.to_numeric(df['ID'], errors='coerce').fillna(0).astype(int)
    df.to_excel(EXCEL_FILE, index=False)

def yeni_id(df: pd.DataFrame) -> int:
    if df.empty:
        return 1
    return int(pd.to_numeric(df['ID'], errors='coerce').max()) + 1

def satir_dogrula(row: dict, idx: int) -> list[str]:
    hatalar = []
    if not str(row.get('Unvan', '')).strip() in GECERLI_UNVANLAR:
        hatalar.append(f"Satır {idx}: Geçersiz Ünvan '{row.get('Unvan')}'")
    if not str(row.get('Ad', '')).strip():
        hatalar.append(f"Satır {idx}: Ad boş olamaz")
    if not str(row.get('Soyad', '')).strip():
        hatalar.append(f"Satır {idx}: Soyad boş olamaz")
    if not str(row.get('Numara_TC', '')).strip():
        hatalar.append(f"Satır {idx}: Numara/TC boş olamaz")
    return hatalar

# ─── RAPORLAR ──────────────────────────────────────────────────────────────────

def excel_rapor_olustur(df: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Siber Panel Raporu"

    # Başlık
    ws.merge_cells('A1:F1')
    ws['A1'] = "⚡ SİBER PANEL PRO — KAYIT RAPORU"
    ws['A1'].font = Font(name='Calibri', bold=True, size=18, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="2A2D3E")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # Tarih
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')} | Toplam Kayıt: {len(df)}"
    ws['A2'].font = Font(name='Calibri', size=10, color="A98BFF")
    ws['A2'].fill = PatternFill("solid", fgColor="1A1D2E")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Tablo başlıkları
    basliklar = ['Sıra', 'ID', 'Ünvan', 'Ad', 'Soyad', 'Numara / TC']
    for col, baslik in enumerate(basliklar, 1):
        hucre = ws.cell(row=4, column=col, value=baslik)
        hucre.font = Font(bold=True, size=11, color="FFFFFF")
        hucre.fill = PatternFill("solid", fgColor="7C5CFC")
        hucre.alignment = Alignment(horizontal='center', vertical='center')

    # Veriler
    for i, row in enumerate(df.to_dict(orient='records'), 1):
        satir = i + 4
        hucreler = [i, row.get('ID', ''), row.get('Unvan', ''), row.get('Ad', ''), row.get('Soyad', ''), row.get('Numara_TC', '')]
        for col, deger in enumerate(hucreler, 1):
            hucre = ws.cell(row=satir, column=col, value=deger)
            hucre.font = Font(size=10, color="E8E8F0")
            hucre.fill = PatternFill("solid", fgColor="0D0F1A" if i % 2 == 0 else "111327")
            hucre.alignment = Alignment(horizontal='center' if col <= 3 else 'left', vertical='center')

    # Sütun genişlikleri
    genislikler = [6, 8, 14, 20, 20, 22]
    for col, genislik in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(col)].width = genislik

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def csv_rapor_olustur(df: pd.DataFrame) -> io.BytesIO:
    buf = io.BytesIO()
    buf.write('\uFEFF'.encode('utf-8'))  # BOM for Turkish chars
    
    writer = csv.writer(buf, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    
    # Başlık
    writer.writerow(['Siber Panel Pro - Kayıt Raporu'])
    writer.writerow([f'Tarih: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}'])
    writer.writerow([f'Toplam Kayıt: {len(df)}'])
    writer.writerow([])  # Boş satır
    
    # Kolon başlıkları
    writer.writerow(['ID', 'Ünvan', 'Ad', 'Soyad', 'Numara / TC'])
    
    # Veriler
    for _, row in df.iterrows():
        writer.writerow([row.get('ID', ''), row.get('Unvan', ''), row.get('Ad', ''), row.get('Soyad', ''), row.get('Numara_TC', '')])
    
    buf.seek(0)
    return buf

def pdf_rapor_olustur(df: pd.DataFrame) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import tempfile
    
    buf = io.BytesIO()
    
    # Türkçe karakter desteği için
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', 'DejaVuSans.ttf'))
        font_name = 'DejaVu'
    except:
        font_name = 'Helvetica'
    
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    
    # Özel stiller
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName=font_name,
        fontSize=18,
        textColor=colors.HexColor('#7d5fff'),
        alignment=1,  # Center
        spaceAfter=20
    )
    
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=10,
        textColor=colors.gray,
        alignment=1,
        spaceAfter=30
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=9,
        textColor=colors.white,
        alignment=1
    )
    
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=8,
        textColor=colors.black,
        alignment=0
    )
    
    story = []
    
    # Başlık
    story.append(Paragraph("⚡ SİBER PANEL PRO - KAYIT RAPORU", title_style))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}", subtitle_style))
    story.append(Paragraph(f"Toplam Kayıt Sayısı: <b>{len(df)}</b>", subtitle_style))
    story.append(Spacer(1, 1*cm))
    
    # Tablo verisi
    data = []
    headers = ['ID', 'Ünvan', 'Ad', 'Soyad', 'Numara / TC']
    data.append([Paragraph(h, header_style) for h in headers])
    
    for _, row in df.iterrows():
        data.append([
            Paragraph(str(row.get('ID', '')), cell_style),
            Paragraph(str(row.get('Unvan', '')), cell_style),
            Paragraph(str(row.get('Ad', '')), cell_style),
            Paragraph(str(row.get('Soyad', '')), cell_style),
            Paragraph(str(row.get('Numara_TC', '')), cell_style),
        ])
    
    # Tablo
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7d5fff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(table)
    doc.build(story)
    buf.seek(0)
    return buf

# ─── ROTALAR ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    df = verileri_oku()
    stats = {
        'toplam': len(df),
        'ogrenci': len(df[df['Unvan'] == 'Öğrenci']) if not df.empty else 0,
        'ogretmen': len(df[df['Unvan'] == 'Öğretmen']) if not df.empty else 0,
        'veli': len(df[df['Unvan'] == 'Veli']) if not df.empty else 0,
    }
    return render_template('index.html', veriler=df.to_dict(orient='records'), stats=stats)

@app.route('/ekle', methods=['GET', 'POST'])
def ekle():
    if request.method == 'POST':
        df = verileri_oku()
        yeni = pd.DataFrame([{
            'ID': yeni_id(df), 'Unvan': request.form.get('unvan'),
            'Ad': request.form.get('ad'), 'Soyad': request.form.get('soyad'),
            'Numara_TC': request.form.get('numara'),
        }])
        df = pd.concat([df, yeni], ignore_index=True)
        verileri_yaz(df)
        flash("✅ Başarıyla eklendi!", "success")
        return redirect(url_for('index'))
    return render_template('ekle.html')

@app.route('/guncelle/<int:id>', methods=['GET', 'POST'])
def guncelle(id):
    df = verileri_oku()
    idx = df[df['ID'] == id].index
    if idx.empty:
        flash("❌ Kayıt bulunamadı!", "danger")
        return redirect(url_for('index'))
    if request.method == 'POST':
        for k, f in [('Unvan', 'unvan'), ('Ad', 'ad'), ('Soyad', 'soyad'), ('Numara_TC', 'numara')]:
            df.at[idx[0], k] = request.form.get(f)
        verileri_yaz(df)
        flash("✅ Kayıt güncellendi!", "info")
        return redirect(url_for('index'))
    return render_template('guncelleme.html', kisi=df.loc[idx[0]].to_dict())

@app.route('/sil/<int:id>', methods=['GET', 'POST'])
def sil(id):
    df = verileri_oku()
    idx = df[df['ID'] == id].index
    if idx.empty:
        flash("❌ Kayıt bulunamadı!", "danger")
        return redirect(url_for('index'))
    kisi = df.loc[idx[0]].to_dict()
    if request.method == 'POST':
        df = df[df['ID'] != id].reset_index(drop=True)
        verileri_yaz(df)
        flash("🗑️ Kayıt silindi!", "danger")
        return redirect(url_for('index'))
    return render_template('sil.html', kisi=kisi)

@app.route('/coklu_sil', methods=['POST'])
def coklu_sil():
    ids = [int(x) for x in request.form.getlist('secili_id')]
    if ids:
        df = verileri_oku()
        df = df[~df['ID'].isin(ids)].reset_index(drop=True)
        verileri_yaz(df)
        flash(f"🗑️ {len(ids)} kayıt silindi!", "danger")
    return redirect(url_for('index'))

@app.route('/temizle', methods=['POST'])
def temizle():
    verileri_yaz(pd.DataFrame(columns=KOLONLAR))
    flash("💀 Tüm liste sıfırlandı!", "danger")
    return redirect(url_for('index'))

@app.route('/excel_yukle', methods=['POST'])
def excel_yukle():
    dosya = request.files.get('file')
    if not dosya or not dosya.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'ok': False, 'mesaj': 'Geçerli bir Excel dosyası seçin!'})
    try:
        yeni_df = pd.read_excel(dosya, dtype=str).fillna("")
        yeni_df.columns = [c.strip() for c in yeni_df.columns]
        zorunlu = {'Unvan', 'Ad', 'Soyad', 'Numara_TC'}
        eksik = zorunlu - set(yeni_df.columns)
        if eksik:
            return jsonify({'ok': False, 'mesaj': f"Eksik sütunlar: {', '.join(eksik)}"})
        
        hatalar = []
        for i, row in enumerate(yeni_df.to_dict(orient='records'), 1):
            hatalar.extend(satir_dogrula(row, i))
        if hatalar:
            return jsonify({'ok': False, 'hatalar': hatalar, 'mesaj': f"{len(hatalar)} hata bulundu."})
        
        mevcut = verileri_oku()
        bas_id = yeni_id(mevcut)
        temiz = yeni_df[['Unvan', 'Ad', 'Soyad', 'Numara_TC']].copy()
        temiz.insert(0, 'ID', range(bas_id, bas_id + len(temiz)))
        birlesik = pd.concat([mevcut, temiz], ignore_index=True)
        verileri_yaz(birlesik)
        return jsonify({'ok': True, 'eklenen': len(temiz), 'mesaj': f"{len(temiz)} kayıt aktarıldı!"})
    except Exception as e:
        return jsonify({'ok': False, 'mesaj': f"Hata: {str(e)}"})

# ─── RAPOR İNDİRME ROTALARI ────────────────────────────────────────────────────

@app.route('/rapor_indir')
def rapor_indir_excel():
    df = verileri_oku()
    buf = excel_rapor_olustur(df)
    tarih = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True, download_name=f'siber_panel_rapor_{tarih}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/rapor_csv')
def rapor_indir_csv():
    df = verileri_oku()
    buf = csv_rapor_olustur(df)
    tarih = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True, download_name=f'siber_panel_rapor_{tarih}.csv', mimetype='text/csv')

@app.route('/rapor_pdf')
def rapor_indir_pdf():
    df = verileri_oku()
    try:
        buf = pdf_rapor_olustur(df)
        tarih = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(buf, as_attachment=True, download_name=f'siber_panel_rapor_{tarih}.pdf', mimetype='application/pdf')
    except Exception as e:
        flash(f"PDF oluşturulamadı: {str(e)}", "danger")
        return redirect(url_for('index'))

if __name__ == '__main__':
    veritabani_kontrol()
    app.run(debug=True, port=5001)